from flask import jsonify, Blueprint, send_file, request
from flask_jwt_extended import jwt_required
from app.models import (
    db,
    Produto,
    Categoria,
    Fornecedor,
    Abastece,
    Pedido,
    ItemPedido,
    Cliente,
    Usuario,
    MovimentacaoEstoque,
    TipoMovimentacao
)
from datetime import datetime, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


dashboard_bp = Blueprint('dashboard', __name__)


def converter_data(valor):
    if not valor:
        return None

    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        return None


def formatar_data(valor):
    if not valor:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    return str(valor)


def aplicar_estilo_planilha(ws, titulo):
    ws.insert_rows(1, 2)

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=16, color="4E1633")

    header_fill = PatternFill("solid", fgColor="F3E6EF")
    header_font = Font(bold=True, color="4E1633")
    border = Border(
        bottom=Side(style="thin", color="C36196")
    )

    header_row = 3

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

    ws.freeze_panes = "A4"


def criar_workbook(nome_aba, titulo, cabecalhos, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    ws.append(cabecalhos)

    for linha in linhas:
        ws.append(linha)

    aplicar_estilo_planilha(ws, titulo)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return arquivo


@dashboard_bp.route('/exportar', methods=['GET'])
@jwt_required()
def exportar_relatorio():
    tipo = request.args.get("tipo", "vendas")
    data_inicio = converter_data(request.args.get("data_inicio"))
    data_fim = converter_data(request.args.get("data_fim"))

    if tipo == "vendas":
        return exportar_vendas(data_inicio, data_fim)

    if tipo == "estoque":
        return exportar_estoque()

    if tipo == "validade":
        return exportar_validade()

    if tipo == "movimentacoes":
        return exportar_movimentacoes(data_inicio, data_fim)

    return jsonify({"erro": "Tipo de relatório inválido."}), 400


def exportar_vendas(data_inicio=None, data_fim=None):
    query = (
        db.session.query(Pedido)
        .join(Cliente, Cliente.id_cliente == Pedido.id_cliente)
        .join(Usuario, Usuario.id_usuario == Pedido.id_usuario)
        .order_by(Pedido.dt_pedido.desc())
    )

    if data_inicio:
        query = query.filter(Pedido.dt_pedido >= datetime.combine(data_inicio, datetime.min.time()))

    if data_fim:
        query = query.filter(Pedido.dt_pedido <= datetime.combine(data_fim, datetime.max.time()))

    pedidos = query.all()

    cabecalhos = [
        "ID Pedido",
        "Data",
        "Cliente",
        "Vendedor",
        "Método de pagamento",
        "Produto",
        "Quantidade",
        "Preço unitário",
        "Subtotal item",
        "Desconto do pedido"
    ]

    linhas = []

    for pedido in pedidos:
        itens = ItemPedido.query.filter_by(id_pedido=pedido.id_pedido).all()

        for item in itens:
            produto = Produto.query.get(item.id_produto)

            preco = float(item.preco_unitario or 0)
            quantidade = int(item.qtdd_pedido or 0)
            subtotal = preco * quantidade

            linhas.append([
                pedido.id_pedido,
                formatar_data(pedido.dt_pedido),
                pedido.cliente.razao_social if pedido.cliente else "",
                pedido.vendedor.nome_usuario if pedido.vendedor else "",
                pedido.mtd_pagamento,
                produto.nome_produto if produto else "",
                quantidade,
                preco,
                subtotal,
                float(pedido.desconto or 0)
            ])

    arquivo = criar_workbook(
        "Vendas",
        "Relatório de Vendas",
        cabecalhos,
        linhas
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_vendas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def exportar_estoque():
    produtos = Produto.query.order_by(Produto.nome_produto.asc()).all()

    cabecalhos = [
        "ID Produto",
        "Produto",
        "Categoria",
        "Sabor",
        "Marca",
        "Estoque atual",
        "Custo unitário",
        "Preço de venda",
        "Valor potencial de venda",
        "Status"
    ]

    linhas = []

    for produto in produtos:
        estoque = int(produto.qtdd_atual or 0)
        custo = float(produto.custo_unitario or 0)
        preco = float(produto.preco_unitario or 0)

        if produto.ativo:
            status = "Ativo"
        else:
            status = "Inativo"

        linhas.append([
            produto.id_produto,
            produto.nome_produto,
            produto.categoria.nome if produto.categoria else "",
            produto.sabor or "",
            produto.marca or "",
            estoque,
            custo,
            preco,
            estoque * preco,
            status
        ])

    arquivo = criar_workbook(
        "Estoque",
        "Relatório de Estoque",
        cabecalhos,
        linhas
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_estoque.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def exportar_validade():
    abastecimentos = (
        Abastece.query
        .filter(Abastece.qtdd_disponivel > 0)
        .order_by(Abastece.data_vencimento.asc())
        .all()
    )

    cabecalhos = [
        "Produto",
        "Lote",
        "Fornecedor",
        "Quantidade recebida",
        "Quantidade disponível",
        "Data de vencimento",
        "Dias restantes",
        "Custo unitário"
    ]

    linhas = []
    hoje = date.today()

    for ab in abastecimentos:
        produto = Produto.query.get(ab.id_produto)
        fornecedor = Fornecedor.query.get(ab.id_fornecedor)

        dias_restantes = ""
        if ab.data_vencimento:
            dias_restantes = (ab.data_vencimento - hoje).days

        linhas.append([
            produto.nome_produto if produto else "",
            ab.numero_lote or "",
            fornecedor.razao_social if fornecedor else "",
            int(ab.qtdd_recebida or 0),
            int(ab.qtdd_disponivel or 0),
            formatar_data(ab.data_vencimento),
            dias_restantes,
            float(ab.valor_unitario or 0)
        ])

    arquivo = criar_workbook(
        "Validade",
        "Relatório de Produtos por Validade",
        cabecalhos,
        linhas
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_validade.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def exportar_movimentacoes(data_inicio=None, data_fim=None):
    query = (
        db.session.query(MovimentacaoEstoque)
        .order_by(MovimentacaoEstoque.ultima_atualizacao.desc())
    )

    if data_inicio:
        query = query.filter(
            MovimentacaoEstoque.ultima_atualizacao >= datetime.combine(data_inicio, datetime.min.time())
        )

    if data_fim:
        query = query.filter(
            MovimentacaoEstoque.ultima_atualizacao <= datetime.combine(data_fim, datetime.max.time())
        )

    movimentacoes = query.all()

    cabecalhos = [
        "Data",
        "Produto",
        "Tipo",
        "Quantidade",
        "Usuário"
    ]

    linhas = []

    for mov in movimentacoes:
        produto = Produto.query.get(mov.id_produto)
        tipo = TipoMovimentacao.query.get(mov.id_tipo_movimentacao)
        usuario = Usuario.query.get(mov.id_usuario)

        linhas.append([
            formatar_data(mov.ultima_atualizacao),
            produto.nome_produto if produto else "",
            tipo.tipo_movimentacao if tipo else "",
            int(mov.qtdd_movimentacao or 0),
            usuario.nome_usuario if usuario else ""
        ])

    arquivo = criar_workbook(
        "Movimentações",
        "Relatório de Movimentações de Estoque",
        cabecalhos,
        linhas
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_movimentacoes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )